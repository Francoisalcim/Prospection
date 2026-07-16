#!/usr/bin/env python3
"""
Clinical Trials Prospector - V3 with Flexible Data Extraction
Allows users to specify AND/OR operators directly in keyword search
"""

import requests
import json
import csv
import time
from typing import List, Dict, Set, Optional, Tuple, Any
from collections import defaultdict
from datetime import datetime
import re


class ClinicalTrialsProspector:
    """Handles fetching and processing clinical trials data with flexible extraction"""
    
    BASE_URL = "https://clinicaltrials.gov/api/v2/studies"

    VALID_STATUSES = {
        "RECRUITING", "ACTIVE_NOT_RECRUITING", "NOT_YET_RECRUITING",
        "COMPLETED", "TERMINATED", "WITHDRAWN", "SUSPENDED",
        "ENROLLING_BY_INVITATION", "UNKNOWN"
    }

    VALID_PHASES = {
        "EARLY_PHASE1", "PHASE1", "PHASE2", "PHASE3", "PHASE4", "NA"
    }
    
    # Organization type categories
    ORGANIZATION_TYPES = {
        'university': {
            'keywords': ['university', 'universite', 'universität', 'universidad', 'universiti', 
                        'college', 'école'],
            'label': 'Universities & Colleges'
        },
        'institute': {
            'keywords': ['institute', 'institut', 'instituto', 'research center', 'research centre'],
            'label': 'Research Institutes'
        },
        'hospital': {
            'keywords': ['hospital', 'medical center', 'medical centre', 'health system', 
                        'health center', 'clinic', 'clinique', 'klinik'],
            'label': 'Hospitals & Medical Centers'
        },
        'government': {
            'keywords': ['national institutes', 'nih', 'ministry of health', 'department of health',
                        'veterans affairs', 'va medical', 'public health', 'government'],
            'label': 'Government Agencies'
        },
        'foundation': {
            'keywords': ['foundation', 'fondation', 'fundacion', 'stichting', 'trust fund'],
            'label': 'Foundations & Trusts'
        },
        'academic': {
            'keywords': ['school of medicine', 'school of pharmacy', 'faculty of', 
                        'academy', 'academie', 'academic'],
            'label': 'Academic Medical Centers'
        },
        'nonprofit': {
            'keywords': ['nonprofit', 'non-profit', 'charity', 'charitable', 'society', 
                        'association', 'organization'],
            'label': 'Non-Profit Organizations'
        },
        'company': {
            'keywords': [],
            'label': 'Commercial Companies'
        }
    }
    
    # Data extraction options - what users can choose to extract
    DATA_EXTRACTION_OPTIONS = {
        'sponsors': {
            'label': 'Sponsors & Collaborators',
            'description': 'Lead sponsors and collaborating organizations',
            'default': True
        },
        'investigators': {
            'label': 'Principal Investigators',
            'description': 'Lead researchers and their affiliations',
            'default': False
        },
        'locations': {
            'label': 'Study Locations',
            'description': 'Facilities, cities, countries where study is conducted',
            'default': False
        },
        'interventions': {
            'label': 'Interventions',
            'description': 'Drugs, devices, procedures being tested',
            'default': False
        },
        'conditions': {
            'label': 'Conditions',
            'description': 'Diseases and conditions being studied',
            'default': False
        },
        'outcomes': {
            'label': 'Study Outcomes',
            'description': 'Primary and secondary outcome measures',
            'default': False
        },
        'design': {
            'label': 'Study Design',
            'description': 'Phase, type, enrollment, randomization details',
            'default': False
        },
        'eligibility': {
            'label': 'Eligibility Criteria',
            'description': 'Age, gender, inclusion/exclusion criteria',
            'default': False
        },
        'contacts': {
            'label': 'Contact Information',
            'description': 'Recruitment contacts and emails',
            'default': False
        },
        'timeline': {
            'label': 'Dates & Timeline',
            'description': 'Start date, completion date, last update',
            'default': False
        }
    }
    
    def __init__(self, include_types=None, exclude_types=None, extraction_options=None):
        """
        Initialize prospector with filtering and extraction options
        
        Args:
            include_types: List of organization types to INCLUDE
            exclude_types: List of organization types to EXCLUDE
            extraction_options: List of data fields to extract (default: ['sponsors'])
        """
        self.trials_data = []
        self.extracted_data = []  # New: flexible extracted data
        self.include_types = include_types
        self.exclude_types = exclude_types or []
        self.extraction_options = extraction_options or ['sponsors']
        self.last_query_term = ""
        self.last_total_count = None
        self.last_request_debug = {"requests": [], "errors": []}
        self.extraction_diagnostics = {
            "input_trials": 0,
            "included_records": 0,
            "excluded_by_org_filter": 0,
            "processing_errors": 0,
            "organization_type_counts": {},
            "selected_organization_types": self.include_types,
            "extraction_options": self.extraction_options
        }
        
    def get_organization_type(self, name: str) -> str:
        """Determine the organization type based on its name"""
        if not name:
            return 'unknown'
            
        name_lower = name.lower()
        
        for org_type, info in self.ORGANIZATION_TYPES.items():
            if org_type == 'company':
                continue
            for keyword in info['keywords']:
                if keyword in name_lower:
                    return org_type
        
        return 'company'
    
    def should_include_organization(self, name: str) -> bool:
        """Check if organization should be included based on filtering rules"""
        if not name:
            return False
        
        org_type = self.get_organization_type(name)
        
        if self.include_types is not None:
            return org_type in self.include_types
        
        return org_type not in self.exclude_types
    
    def _split_top_level_commas(self, text: str) -> List[str]:
        """
        Split on commas, but not commas inside parentheses.

        Example:
        'Acne, Chronic wounds (diabetic, venous ulcers), Xerosis'
        ->
        ['Acne', 'Chronic wounds (diabetic, venous ulcers)', 'Xerosis']
        """
        parts = []
        current = []
        depth = 0

        for char in text:
            if char == "(":
                depth += 1
                current.append(char)
            elif char == ")":
                depth = max(0, depth - 1)
                current.append(char)
            elif char == "," and depth == 0:
                part = "".join(current).strip()
                if part:
                    parts.append(part)
                current = []
            else:
                current.append(char)

        part = "".join(current).strip()
        if part:
            parts.append(part)

        return parts


    def parse_keyword_expression(self, keyword_string: str) -> str:
        """
        Parse user keyword input safely.

        Supported:
        - acne
        - acne, alopecia, xerosis
        - chronic wounds (diabetic, venous ulcers), acne
        - diabetes AND insulin
        - cancer AND (immunotherapy OR chemotherapy)

        Important:
        Parentheses inside disease names are allowed in comma-separated searches.
        Parentheses are treated as Boolean grouping only when AND/OR is used.
        """
        keyword_string = keyword_string.strip()

        if not keyword_string:
            return ""

        keyword_string = re.sub(r"\s+", " ", keyword_string)

        has_boolean = re.search(r"\bAND\b|\bOR\b", keyword_string, flags=re.IGNORECASE)

        # Case 1 — comma-separated disease/condition list
        if not has_boolean:
            terms = self._split_top_level_commas(keyword_string)

            if len(terms) == 1:
                return terms[0]

            # Quote each term so multi-word medical conditions stay together
            quoted_terms = []
            for term in terms:
                clean_term = term.strip()
                if clean_term.startswith('"') and clean_term.endswith('"'):
                    quoted_terms.append(clean_term)
                else:
                    quoted_terms.append(f'"{clean_term}"')

            return "(" + " OR ".join(quoted_terms) + ")"

        # Case 2 — explicit Boolean search
        expression = re.sub(r"\bAND\b", "AND", keyword_string, flags=re.IGNORECASE)
        expression = re.sub(r"\bOR\b", "OR", expression, flags=re.IGNORECASE)

        # Validate parentheses only for Boolean expressions
        balance = 0
        for char in expression:
            if char == "(":
                balance += 1
            elif char == ")":
                balance -= 1

            if balance < 0:
                raise ValueError(
                    "Invalid keyword expression: closing parenthesis before opening parenthesis."
                )

        if balance != 0:
            raise ValueError("Invalid keyword expression: unbalanced parentheses.")

        if re.search(r"\b(AND|OR)\s+(AND|OR)\b", expression):
            raise ValueError(
                "Invalid keyword expression: two Boolean operators are next to each other."
            )

        if re.search(r"^\s*(AND|OR)\b", expression) or re.search(r"\b(AND|OR)\s*$", expression):
            raise ValueError(
                "Invalid keyword expression: expression cannot start or end with AND/OR."
            )

        return expression

    def _validate_keyword_expression(self, expression: str) -> None:
        """Validate Boolean expression shape before sending it to ClinicalTrials.gov."""
        balance = 0
        previous_token = None
        tokens = re.findall(r"\(|\)|\bAND\b|\bOR\b|[^\s()]+", expression)

        if not tokens:
            raise ValueError("Invalid keyword expression: empty expression.")

        for token in tokens:
            upper = token.upper()

            if token == "(":
                balance += 1
                if previous_token and previous_token not in {"AND", "OR", "("}:
                    raise ValueError("Invalid keyword expression: missing AND/OR before '('.")

            elif token == ")":
                balance -= 1
                if balance < 0:
                    raise ValueError("Invalid keyword expression: closing parenthesis before opening parenthesis.")
                if previous_token in {"AND", "OR", "("}:
                    raise ValueError("Invalid keyword expression: empty group or operator before ')'.")

            elif upper in {"AND", "OR"}:
                if previous_token is None or previous_token in {"AND", "OR", "("}:
                    raise ValueError("Invalid keyword expression: AND/OR must be placed between two terms.")

            else:
                if previous_token == ")":
                    raise ValueError("Invalid keyword expression: missing AND/OR after ')'.")

            previous_token = upper if upper in {"AND", "OR"} else token

        if balance != 0:
            raise ValueError("Invalid keyword expression: unbalanced parentheses.")

        if previous_token in {"AND", "OR", "("}:
            raise ValueError("Invalid keyword expression: expression cannot end with AND/OR or '('.")

    def validate_filters(self, statuses: List[str] = None, phases: List[str] = None) -> None:
        """Validate API filter values before sending the request."""
        if statuses:
            invalid = sorted(set(statuses) - self.VALID_STATUSES)
            if invalid:
                raise ValueError("Invalid status value(s): " + ", ".join(invalid))

        if phases:
            invalid = sorted(set(phases) - self.VALID_PHASES)
            if invalid:
                raise ValueError("Invalid phase value(s): " + ", ".join(invalid))

    def build_query_term(self, keywords: str) -> str:
        """
        Build the query.term parameter for the API
        
        Args:
            keywords: Keyword string (can include AND/OR operators)
            
        Returns:
            Formatted query string for API
        """
        # Only process keywords - phases will be handled as filter parameters
        return self.parse_keyword_expression(keywords)
    
    def fetch_trials(self, 
                     keywords: str,
                     statuses: List[str] = None,
                     phases: List[str] = None,
                     max_results: Optional[int] = 500,
                     progress_callback=None) -> List[Dict]:
        """Fetch clinical trials from ClinicalTrials.gov API with pagination"""
        self.trials_data = []
        self.last_request_debug = {"requests": [], "errors": []}
        self.last_total_count = None
        self.validate_filters(statuses=statuses, phases=phases)
        page_size = 100
        page_token = None
        
        query_term = self.build_query_term(keywords)
        self.last_query_term = query_term
        
        print("🔍 Searching for: {}".format(query_term))
        print("📊 Filters: Statuses={}, Phases={}, Max Results={}".format(
            statuses or 'All', phases or 'All', max_results if max_results is not None else 'ALL'))
        print("📋 Extracting: {}".format(', '.join(self.extraction_options)))
        
        while max_results is None or len(self.trials_data) < max_results:
            params = {
                'query.term': query_term,
                'pageSize': page_size if max_results is None else min(page_size, max_results - len(self.trials_data)),
                'format': 'json',
                'countTotal': 'true'
            }
            
            # Apply status filter if specified
            if statuses:
                params['filter.overallStatus'] = ','.join(statuses)
                       
            if page_token:
                params['pageToken'] = page_token
            
            print("📥 Fetching... ({}/{})".format(len(self.trials_data), max_results if max_results is not None else 'ALL'))
            
            try:
                response = requests.get(self.BASE_URL, params=params, timeout=30)
                request_debug = {
                    'url': response.url,
                    'status_code': response.status_code,
                    'params': params.copy()
                }
                if response.status_code != 200:
                    request_debug['response_preview'] = response.text[:1000]
                    self.last_request_debug['errors'].append(request_debug)
                self.last_request_debug['requests'].append(request_debug)
                response.raise_for_status()
                data = response.json()
                self.last_total_count = data.get('totalCount', self.last_total_count)
                
                studies = data.get('studies', [])
                if phases:
                    requested_phases = set(phases)

                    studies = [
                        study for study in studies
                        if requested_phases.intersection(
                            set(
                                study.get("protocolSection", {})
                                .get("designModule", {})
                                .get("phases", [])
                            )
                        )
                    ]

                if not studies:
                    break
                
                self.trials_data.extend(studies)
                
                if progress_callback:
                    progress_callback({
                        'stage': 'fetching',
                        'fetched': len(self.trials_data),
                        'apiTotalCount': self.last_total_count,
                        'latestPageSize': len(studies)
                    })
                
                page_token = data.get('nextPageToken')
                if not page_token:
                    break
                
                time.sleep(0.2)
                
            except requests.exceptions.RequestException as e:
                print("❌ Error fetching data: {}".format(e))
                raise
        
        print("✅ Fetched {} trials".format(len(self.trials_data)))
        return self.trials_data
    
    def extract_data(self, trials: List[Dict] = None) -> List[Dict]:
        """
        Extract selected data fields from trials
        
        Returns:
            List of dictionaries with extracted data based on user selection
        """
        if trials is None:
            trials = self.trials_data
        
        self.extracted_data = []
        self.extraction_diagnostics = {
            "input_trials": len(trials),
            "included_records": 0,
            "excluded_by_org_filter": 0,
            "processing_errors": 0,
            "organization_type_counts": {},
            "selected_organization_types": self.include_types,
            "extraction_options": self.extraction_options
        }
        
        for study in trials:
            try:
                ps = study.get('protocolSection', {})
                nct_id = ps.get('identificationModule', {}).get('nctId', '')
                
                # Start with base data (always included)
                extracted = {
                    'nct_id': nct_id,
                    'title': ps.get('identificationModule', {}).get('briefTitle', ''),
                    'status': ps.get('statusModule', {}).get('overallStatus', '')
                }
                
                # Extract based on user selections
                if 'sponsors' in self.extraction_options:
                    extracted.update(self._extract_sponsors(ps))
                
                if 'investigators' in self.extraction_options:
                    extracted.update(self._extract_investigators(ps))
                
                if 'locations' in self.extraction_options:
                    extracted.update(self._extract_locations(ps))
                
                if 'interventions' in self.extraction_options:
                    extracted.update(self._extract_interventions(ps))
                
                if 'conditions' in self.extraction_options:
                    extracted.update(self._extract_conditions(ps))
                
                if 'outcomes' in self.extraction_options:
                    extracted.update(self._extract_outcomes(ps))
                
                if 'design' in self.extraction_options:
                    extracted.update(self._extract_design(ps))
                
                if 'eligibility' in self.extraction_options:
                    extracted.update(self._extract_eligibility(ps))
                
                if 'contacts' in self.extraction_options:
                    extracted.update(self._extract_contacts(ps))
                
                if 'timeline' in self.extraction_options:
                    extracted.update(self._extract_timeline(ps))
                
                # Apply organization filtering (if sponsors is selected)
                if 'sponsors' in self.extraction_options:
                    lead_sponsor = extracted.get('lead_sponsor', '')
                    lead_sponsor_type = extracted.get('lead_sponsor_type') or self.get_organization_type(lead_sponsor)
                    counts = self.extraction_diagnostics["organization_type_counts"]
                    counts[lead_sponsor_type] = counts.get(lead_sponsor_type, 0) + 1

                    if lead_sponsor and self.should_include_organization(lead_sponsor):
                        self.extracted_data.append(extracted)
                    else:
                        self.extraction_diagnostics["excluded_by_org_filter"] += 1
                else:
                    # If sponsors are not extracted, org filtering cannot be applied safely.
                    self.extracted_data.append(extracted)
                    
            except Exception as e:
                self.extraction_diagnostics["processing_errors"] += 1
                print("⚠️  Warning: Error processing study {}: {}".format(nct_id, e))
                continue
        
        self.extraction_diagnostics["included_records"] = len(self.extracted_data)
        print(f"\n✅ Extracted data from {len(self.extracted_data)} trials")
        return self.extracted_data
    
    def _extract_sponsors(self, ps: Dict) -> Dict:
        """Extract sponsor and collaborator information"""
        sponsors_module = (ps.get('sponsorsCollaboratorsModule') or 
                          ps.get('sponsorCollaboratorsModule', {}))
        
        lead_sponsor = sponsors_module.get('leadSponsor', {}).get('name', '')
        lead_sponsor_class = sponsors_module.get('leadSponsor', {}).get('class', '')
        
        collaborators = sponsors_module.get('collaborators', [])
        collab_names = [c.get('name', '') for c in collaborators if c.get('name')]
        
        return {
            'lead_sponsor': lead_sponsor,
            'lead_sponsor_class': lead_sponsor_class,
            'lead_sponsor_type': self.get_organization_type(lead_sponsor),
            'collaborators': '; '.join(collab_names),
            'collaborator_count': len(collab_names)
        }
    
    def _extract_investigators(self, ps: Dict) -> Dict:
        """Extract principal investigator information"""
        contacts_module = ps.get('contactsLocationsModule', {})
        officials = contacts_module.get('overallOfficials', [])
        
        pi_names = []
        pi_affiliations = []
        
        for official in officials:
            if official.get('role') in ['PRINCIPAL_INVESTIGATOR', 'STUDY_DIRECTOR']:
                pi_names.append(official.get('name', ''))
                pi_affiliations.append(official.get('affiliation', ''))
        
        return {
            'principal_investigators': '; '.join(pi_names),
            'pi_affiliations': '; '.join(pi_affiliations),
            'pi_count': len(pi_names)
        }
    
    def _extract_locations(self, ps: Dict) -> Dict:
        """Extract study location information"""
        contacts_module = ps.get('contactsLocationsModule', {})
        locations = contacts_module.get('locations', [])
        
        facilities = []
        cities = []
        countries = []
        
        for loc in locations:
            if loc.get('facility'):
                facilities.append(loc['facility'])
            if loc.get('city'):
                cities.append(loc['city'])
            if loc.get('country'):
                countries.append(loc['country'])
        
        return {
            'facilities': '; '.join(set(facilities)),
            'cities': '; '.join(set(cities)),
            'countries': '; '.join(set(countries)),
            'location_count': len(locations)
        }
    
    def _extract_interventions(self, ps: Dict) -> Dict:
        """Extract intervention information"""
        arms_module = ps.get('armsInterventionsModule', {})
        interventions = arms_module.get('interventions', [])
        
        drugs = []
        devices = []
        procedures = []
        other = []
        
        for intervention in interventions:
            int_type = intervention.get('type', '')
            name = intervention.get('name', '')
            
            if int_type == 'DRUG':
                drugs.append(name)
            elif int_type == 'DEVICE':
                devices.append(name)
            elif int_type in ['PROCEDURE', 'SURGERY']:
                procedures.append(name)
            else:
                other.append(name)
        
        return {
            'drugs': '; '.join(drugs),
            'devices': '; '.join(devices),
            'procedures': '; '.join(procedures),
            'other_interventions': '; '.join(other),
            'intervention_count': len(interventions)
        }
    
    def _extract_conditions(self, ps: Dict) -> Dict:
        """Extract condition/disease information"""
        cond_module = ps.get('conditionsModule', {})
        conditions = cond_module.get('conditions', [])
        keywords = cond_module.get('keywords', [])
        
        return {
            'conditions': '; '.join(conditions),
            'keywords': '; '.join(keywords),
            'condition_count': len(conditions)
        }
    
    def _extract_outcomes(self, ps: Dict) -> Dict:
        """Extract outcome measure information"""
        outcomes_module = ps.get('outcomesModule', {})
        primary_outcomes = outcomes_module.get('primaryOutcomes', [])
        secondary_outcomes = outcomes_module.get('secondaryOutcomes', [])
        
        primary_measures = [o.get('measure', '') for o in primary_outcomes]
        secondary_measures = [o.get('measure', '') for o in secondary_outcomes]
        
        return {
            'primary_outcomes': '; '.join(primary_measures),
            'secondary_outcomes': '; '.join(secondary_measures),
            'primary_outcome_count': len(primary_measures),
            'secondary_outcome_count': len(secondary_measures)
        }
    
    def _extract_design(self, ps: Dict) -> Dict:
        """Extract study design information"""
        design_module = ps.get('designModule', {})
        
        phases = design_module.get('phases', [])
        study_type = design_module.get('studyType', '')
        enrollment = design_module.get('enrollmentInfo', {})
        
        design_info = design_module.get('designInfo', {})
        allocation = design_info.get('allocation', '')
        intervention_model = design_info.get('interventionModel', '')
        primary_purpose = design_info.get('primaryPurpose', '')
        masking = design_info.get('maskingInfo', {}).get('masking', '')
        
        return {
            'phase': ', '.join(phases),
            'study_type': study_type,
            'enrollment': enrollment.get('count', 0),
            'allocation': allocation,
            'intervention_model': intervention_model,
            'primary_purpose': primary_purpose,
            'masking': masking
        }
    
    def _extract_eligibility(self, ps: Dict) -> Dict:
        """Extract eligibility criteria"""
        eligibility_module = ps.get('eligibilityModule', {})
        
        return {
            'min_age': eligibility_module.get('minimumAge', ''),
            'max_age': eligibility_module.get('maximumAge', ''),
            'sex': eligibility_module.get('sex', ''),
            'healthy_volunteers': eligibility_module.get('healthyVolunteers', ''),
            'eligibility_criteria': eligibility_module.get('eligibilityCriteria', '')[:500]  # Truncate
        }
    
    def _extract_contacts(self, ps: Dict) -> Dict:
        """Extract contact information"""
        contacts_module = ps.get('contactsLocationsModule', {})
        central_contacts = contacts_module.get('centralContacts', [])
        
        contact_names = []
        contact_emails = []
        contact_phones = []
        
        for contact in central_contacts:
            if contact.get('name'):
                contact_names.append(contact['name'])
            if contact.get('email'):
                contact_emails.append(contact['email'])
            if contact.get('phone'):
                contact_phones.append(contact['phone'])
        
        return {
            'contact_name': '; '.join(contact_names),
            'contact_email': '; '.join(contact_emails),
            'contact_phone': '; '.join(contact_phones)
        }
    
    def _extract_timeline(self, ps: Dict) -> Dict:
        """Extract date and timeline information"""
        status_module = ps.get('statusModule', {})
        
        start_date = status_module.get('startDateStruct', {}).get('date', '')
        completion_date = status_module.get('completionDateStruct', {}).get('date', '')
        last_update = status_module.get('lastUpdatePostDateStruct', {}).get('date', '')
        
        return {
            'start_date': start_date,
            'completion_date': completion_date,
            'last_update': last_update
        }
    
    def export_to_xlsx(self, filename: str = None, column_order: List[str] = None) -> str:
        """Export extracted data to Excel (XLSX) with custom column order"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("❌ openpyxl not installed. Run: pip install openpyxl")
            return None
        
        if not self.extracted_data:
            print("❌ No data to export")
            return None
        
        if filename is None:
            date = datetime.now().strftime('%Y-%m-%d')
            filename = f"ClinicalTrials_Export_{date}.xlsx"
        
        # Get all unique keys from extracted data
        all_keys = set()
        for row in self.extracted_data:
            all_keys.update(row.keys())
        
        # Use custom column order if provided, otherwise use default
        if column_order:
            # Filter to only include columns that actually exist in the data
            fieldnames = [col for col in column_order if col in all_keys]
            # Add any remaining columns that weren't in the custom order
            remaining = [col for col in sorted(all_keys) if col not in fieldnames]
            fieldnames.extend(remaining)
        else:
            # Default behavior: sort keys and prioritize important fields
            fieldnames = sorted(all_keys)
            priority_fields = ['nct_id', 'title', 'status', 'lead_sponsor']
            for field in reversed(priority_fields):
                if field in fieldnames:
                    fieldnames.remove(field)
                    fieldnames.insert(0, field)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clinical Trials Data"
        
        # Style for header
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write header
        for col_idx, fieldname in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = fieldname.replace('_', ' ').title()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, record in enumerate(self.extracted_data, 2):
            for col_idx, fieldname in enumerate(fieldnames, 1):
                value = record.get(fieldname, '')
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # Wrap text for long fields
                if fieldname in ['title', 'eligibility_criteria', 'conditions', 'interventions']:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Auto-adjust column widths
        for col_idx, fieldname in enumerate(fieldnames, 1):
            column_letter = get_column_letter(col_idx)
            
            # Set specific widths for known columns
            if fieldname == 'nct_id':
                ws.column_dimensions[column_letter].width = 12
            elif fieldname == 'title':
                ws.column_dimensions[column_letter].width = 50
            elif fieldname in ['eligibility_criteria', 'conditions', 'interventions', 'collaborators']:
                ws.column_dimensions[column_letter].width = 40
            elif fieldname in ['lead_sponsor', 'principal_investigators', 'facilities']:
                ws.column_dimensions[column_letter].width = 30
            else:
                ws.column_dimensions[column_letter].width = 15
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save workbook
        wb.save(filename)
        
        print("✅ Exported to {} with {} columns in custom order".format(filename, len(fieldnames)))
        return filename
    
    def export_to_csv(self, filename: str = None) -> str:
        """Export extracted data to CSV (legacy method)"""
        if not self.extracted_data:
            print("❌ No data to export")
            return None
        
        if filename is None:
            date = datetime.now().strftime('%Y-%m-%d')
            filename = f"ClinicalTrials_Export_{date}.csv"
        
        # Get all unique keys from extracted data
        all_keys = set()
        for row in self.extracted_data:
            all_keys.update(row.keys())
        
        # Sort keys for consistent column order
        fieldnames = sorted(all_keys)
        
        # Move important fields to front
        priority_fields = ['nct_id', 'title', 'status', 'lead_sponsor']
        for field in reversed(priority_fields):
            if field in fieldnames:
                fieldnames.remove(field)
                fieldnames.insert(0, field)
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.extracted_data)
        
        print("✅ Exported to {}".format(filename))
        return filename


def main():
    """Example usage with keyword expression parsing"""
    
    # Example 1: Simple OR search (comma-separated)
    print("\n" + "="*70)
    print("EXAMPLE 1: Simple OR search (comma-separated)")
    print("="*70)
    prospector = ClinicalTrialsProspector(
        include_types=['company'],
        extraction_options=['sponsors']
    )
    trials = prospector.fetch_trials(keywords='diabetes, insulin', max_results=50)
    data = prospector.extract_data()
    print(f"Extracted {len(data)} records")
    
    # Example 2: AND search
    print("\n" + "="*70)
    print("EXAMPLE 2: AND search - both keywords required")
    print("="*70)
    prospector2 = ClinicalTrialsProspector(
        include_types=['company', 'university'],
        extraction_options=['sponsors', 'investigators']
    )
    trials2 = prospector2.fetch_trials(keywords='cancer AND immunotherapy', max_results=50)
    data2 = prospector2.extract_data()
    print(f"Extracted {len(data2)} records")
    
    # Example 3: Complex expression with parentheses
    print("\n" + "="*70)
    print("EXAMPLE 3: Complex expression")
    print("="*70)
    prospector3 = ClinicalTrialsProspector(
        include_types=['company'],
        extraction_options=['sponsors', 'design']
    )
    trials3 = prospector3.fetch_trials(
        keywords='diabetes AND (insulin OR metformin)', 
        max_results=50
    )
    data3 = prospector3.extract_data()
    prospector3.export_to_csv('complex_search.csv')


if __name__ == "__main__":
    main()
